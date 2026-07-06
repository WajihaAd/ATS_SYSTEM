# """
# exporter.py
# """

# import io
# import re
# from datetime import datetime

# import pandas as pd
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.utils import get_column_letter


# # ------------------------------------------------------------------
# # REMOVE ILLEGAL EXCEL CHARACTERS
# # ------------------------------------------------------------------

# ILLEGAL_CHARACTERS_RE = re.compile(
#     r'[\000-\010]|[\013-\014]|[\016-\037]'
# )

# def clean_excel_text(value):
#     """
#     Remove characters that Excel/OpenPyXL cannot handle.
#     """
#     if pd.isna(value):
#         return value

#     value = str(value)
#     value = ILLEGAL_CHARACTERS_RE.sub("", value)

#     # Remove other problematic unicode chars
#     value = value.replace("\x00", "")
#     value = value.replace("\ufffd", "")

#     return value


# def clean_dataframe(df):
#     """
#     Clean every string cell before export.
#     Compatible with pandas 2.x
#     """
#     if df is None or df.empty:
#         return df

#     # apply column-wise cleaning
#     return df.apply(lambda col: col.apply(clean_excel_text))


# # ------------------------------------------------------------------
# # STYLING
# # ------------------------------------------------------------------

# def _style_header(ws, row: int, n_cols: int, color: str = "1F4E79"):
#     fill = PatternFill(
#         start_color=color,
#         end_color=color,
#         fill_type="solid"
#     )

#     font = Font(
#         color="FFFFFF",
#         bold=True,
#         size=11
#     )

#     for c in range(1, n_cols + 1):
#         cell = ws.cell(row=row, column=c)
#         cell.fill = fill
#         cell.font = font
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center",
#             wrap_text=True
#         )


# def _auto_fit(ws, max_w: int = 50):
#     for col in ws.columns:
#         letter = get_column_letter(col[0].column)

#         width = max(
#             (len(str(cell.value or "")) for cell in col),
#             default=8
#         )

#         ws.column_dimensions[letter].width = min(
#             width + 2,
#             max_w
#         )


# def _score_fill(score: float):

#     if score >= 70:
#         return PatternFill(
#             start_color="C6EFCE",
#             end_color="C6EFCE",
#             fill_type="solid"
#         )

#     elif score >= 45:
#         return PatternFill(
#             start_color="FFEB9C",
#             end_color="FFEB9C",
#             fill_type="solid"
#         )

#     return PatternFill(
#         start_color="FFC7CE",
#         end_color="FFC7CE",
#         fill_type="solid"
#     )


# def _apply_score_coloring(
#     ws,
#     df: pd.DataFrame,
#     col_name: str = "ATS Score"
# ):

#     if df.empty:
#         return

#     headers = [
#         ws.cell(1, c).value
#         for c in range(1, ws.max_column + 1)
#     ]

#     score_col = next(
#         (
#             i + 1
#             for i, h in enumerate(headers)
#             if h == col_name
#         ),
#         None
#     )

#     if score_col:

#         for row in range(2, ws.max_row + 1):

#             cell = ws.cell(
#                 row=row,
#                 column=score_col
#             )

#             try:
#                 cell.fill = _score_fill(
#                     float(cell.value or 0)
#                 )
#             except Exception:
#                 pass


# # ------------------------------------------------------------------
# # MAIN EXPORTER
# # ------------------------------------------------------------------

# def generate_excel(pipeline_output: dict) -> bytes:

#     jd = pipeline_output.get("jd", {})

#     df_all = clean_dataframe(
#         pipeline_output.get("df_all", pd.DataFrame())
#     )

#     df_top10 = clean_dataframe(
#         pipeline_output.get("df_top10", pd.DataFrame())
#     )

#     df_top20 = clean_dataframe(
#         pipeline_output.get("df_top20", pd.DataFrame())
#     )

#     buf = io.BytesIO()

#     with pd.ExcelWriter(
#         buf,
#         engine="openpyxl"
#     ) as writer:

#         # --------------------------------------------------
#         # SUMMARY
#         # --------------------------------------------------

#         summary = {
#             "Metric": [
#                 "Report Generated",
#                 "Department",
#                 "Seniority Level",
#                 "Required Degree",
#                 "Minimum CGPA",
#                 "Required Experience",
#                 "Required Skills Count",
#                 "Total Candidates Ranked",
#                 "Top Score",
#                 "Average Score",
#             ],

#             "Value": [
#                 datetime.now().strftime("%Y-%m-%d %H:%M"),
#                 jd.get("department", "N/A"),
#                 jd.get("seniority", "N/A"),
#                 jd.get("required_degree", "N/A"),
#                 jd.get("required_cgpa", "N/A"),
#                 jd.get("required_experience_years", 0),
#                 len(jd.get("required_skills", [])),
#                 len(df_all),
#                 df_all["ATS Score"].max()
#                 if not df_all.empty else "N/A",
#                 round(df_all["ATS Score"].mean(), 1)
#                 if not df_all.empty else "N/A",
#             ],
#         }

#         pd.DataFrame(summary).to_excel(
#             writer,
#             sheet_name="Summary",
#             index=False
#         )

#         # --------------------------------------------------
#         # ALL CANDIDATES
#         # --------------------------------------------------

#         if not df_all.empty:
#             df_all.to_excel(
#                 writer,
#                 sheet_name="All Candidates",
#                 index=False
#             )

#         # --------------------------------------------------
#         # TOP 10
#         # --------------------------------------------------

#         if not df_top10.empty:
#             df_top10.to_excel(
#                 writer,
#                 sheet_name="Top 10",
#                 index=False
#             )

#         # --------------------------------------------------
#         # TOP 20
#         # --------------------------------------------------

#         if not df_top20.empty:
#             df_top20.to_excel(
#                 writer,
#                 sheet_name="Top 20",
#                 index=False
#             )

#         # --------------------------------------------------
#         # WEIGHTS
#         # --------------------------------------------------

#         weights_df = pd.DataFrame([
#             {
#                 "Component": k,
#                 "Weight (%)": v
#             }
#             for k, v in jd.get(
#                 "weights",
#                 {}
#             ).items()
#         ])

#         weights_df.to_excel(
#             writer,
#             sheet_name="ATS Weights",
#             index=False
#         )

#         # --------------------------------------------------
#         # REQUIRED SKILLS
#         # --------------------------------------------------

#         skills_df = pd.DataFrame({
#             "Required Skills":
#             jd.get("required_skills", [])
#         })

#         skills_df = clean_dataframe(skills_df)

#         skills_df.to_excel(
#             writer,
#             sheet_name="Required Skills",
#             index=False
#         )

#         # --------------------------------------------------
#         # STYLING
#         # --------------------------------------------------

#         wb = writer.book

#         for name in wb.sheetnames:

#             ws = wb[name]

#             _style_header(
#                 ws,
#                 1,
#                 ws.max_column
#             )

#             _auto_fit(ws)

#             ws.freeze_panes = "A2"

#         if "All Candidates" in wb.sheetnames:
#             _apply_score_coloring(
#                 wb["All Candidates"],
#                 df_all
#             )

#         if "Top 10" in wb.sheetnames:
#             _apply_score_coloring(
#                 wb["Top 10"],
#                 df_top10
#             )

#         if "Top 20" in wb.sheetnames:
#             _apply_score_coloring(
#                 wb["Top 20"],
#                 df_top20
#             )

#     return buf.getvalue()


# # ------------------------------------------------------------------
# # CSV EXPORTS
# # ------------------------------------------------------------------

# def generate_csv_all(pipeline_output):
#     return clean_dataframe(
#         pipeline_output.get(
#             "df_all",
#             pd.DataFrame()
#         )
#     ).to_csv(index=False)


# def generate_csv_top10(pipeline_output):
#     return clean_dataframe(
#         pipeline_output.get(
#             "df_top10",
#             pd.DataFrame()
#         )
#     ).to_csv(index=False)


# def generate_csv_top20(pipeline_output):
#     return clean_dataframe(
#         pipeline_output.get(
#             "df_top20",
#             pd.DataFrame()
#         )
#     ).to_csv(index=False)
"""
exporter.py — FIXED (Pandas 2.x + Excel safe export)
"""

import io
import re
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── SAFE TEXT CLEANER ─────────────────────────────────────────────────────────
ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def clean_excel_text(v):
    if v is None:
        return ""

    if isinstance(v, (int, float)):
        return v

    v = str(v)
    v = ILLEGAL.sub("", v)
    v = v.replace("\ufffd", "")
    return v.strip()


# ── FIXED FOR PANDAS 2.x ──────────────────────────────────────────────────────
def clean_dataframe(df):
    if df is None or df.empty:
        return df

    return df.apply(lambda col: col.map(clean_excel_text))


# ── MAIN EXCEL EXPORT ─────────────────────────────────────────────────────────
def generate_excel(output: dict):

    df_all = clean_dataframe(output.get("df_all", pd.DataFrame()))
    df_top10 = clean_dataframe(output.get("df_top10", pd.DataFrame()))
    df_top20 = clean_dataframe(output.get("df_top20", pd.DataFrame()))
    jd = output.get("jd", {})

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        # SUMMARY
        summary = pd.DataFrame({
            "Metric": ["Total", "Top Score", "Avg Score"],
            "Value": [
                len(df_all),
                df_all["ATS Score"].max() if not df_all.empty else 0,
                df_all["ATS Score"].mean() if not df_all.empty else 0
            ]
        })

        summary.to_excel(writer, sheet_name="Summary", index=False)

        df_all.to_excel(writer, sheet_name="All", index=False)
        df_top10.to_excel(writer, sheet_name="Top10", index=False)
        df_top20.to_excel(writer, sheet_name="Top20", index=False)

    return buffer.getvalue()


# ── CSV EXPORTS ───────────────────────────────────────────────────────────────
def generate_csv_all(out):
    return clean_dataframe(out.get("df_all", pd.DataFrame())).to_csv(index=False)

def generate_csv_top10(out):
    return clean_dataframe(out.get("df_top10", pd.DataFrame())).to_csv(index=False)

def generate_csv_top20(out):
    return clean_dataframe(out.get("df_top20", pd.DataFrame())).to_csv(index=False)